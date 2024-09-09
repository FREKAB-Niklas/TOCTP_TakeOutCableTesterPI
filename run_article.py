import time
import tkinter as tk
from tkinter import font, messagebox
from PIL import Image, ImageTk
import pandas as pd
import os
import sys
import traceback
import subprocess
from datetime import datetime, timedelta
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as OpenpyxlImage
import configparser
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill
import threading
from adafruit_mcp230xx.mcp23017 import MCP23017
import board
import busio
from digitalio import Direction, Pull
import math
import paho.mqtt.client as mqtt
import shutil
import configparser


print(f"{datetime.now()}: run_article.py is starting...")

# Static cable diameter
CABLE_DIAMETER = 8.0  # in mm

# Add this with your other global variables at the top of the script
current_serial_number = 0

# MQTT setup
mqtt_broker_ip = "192.168.10.9"  # Replace with your actual broker IP
client = mqtt.Client()
client.connect(mqtt_broker_ip, 1883, 60)

# Initialize the current segment index
current_segment = 0
rotation_list = []


# Get the directory where the script is located
script_dir = os.path.dirname(os.path.abspath(__file__))
success_sound_path = os.path.join(script_dir, "success.mp3")
reject_sound_path = os.path.join(script_dir, "reject.mp3")

# Set environment variables for SDL to use the PulseAudio driver
os.environ["SDL_AUDIODRIVER"] = "pulseaudio"
os.environ["AUDIODEV"] = "hdmi:CARD=vc4hdmi1"  # Replace "hw:0,0" with the correct device from aplay --list-pcms

import pygame

# Function to start PulseAudio if it's not running
def start_pulseaudio():
    try:
        # Check if PulseAudio is running
        subprocess.run(["pulseaudio", "--check"], check=True)
        print(f"{datetime.now()}: PulseAudio is already running.")
    except subprocess.CalledProcessError:
        # Start PulseAudio if not running
        subprocess.run(["pulseaudio", "--start"])
        print(f"{datetime.now()}: PulseAudio started.")

# Attempt to initialize pygame for sound effects
pygame_initialized = False

# Start PulseAudio
start_pulseaudio()

for _ in range(5):
    try:
        pygame.mixer.init()
        success_sound = pygame.mixer.Sound("success.mp3")
        reject_sound = pygame.mixer.Sound("reject.mp3")
        success_sound.set_volume(1.0)
        reject_sound.set_volume(1.0)
        print(f"{datetime.now()}: Pygame initialized.")
        pygame_initialized = True
        sound = pygame.mixer.Sound("success.mp3")
        break
    except pygame.error as e:
        print(f"{datetime.now()}: Pygame initialization failed: {e}. Retrying...")
        time.sleep(2)  # Wait for 2 seconds before retrying

try:
    success_sound.play()
    pygame.time.wait(3000)
except Exception as e:
    print(f"Error playing success sound: {e}")

if not pygame_initialized:
    print(f"{datetime.now()}: Pygame initialization failed after retries. Continuing without sound.")
    success_sound = None
    reject_sound = None
else:
    print(f"{datetime.now()}: Pygame initialized successfully with audio support.")


def read_last_serial_number_from_log(filename):
    try:
        wb = openpyxl.load_workbook(filename)
        ws_main = wb['Main']
        last_row = ws_main.max_row
        total_count = ws_main['G4'].value
        return total_count if total_count else 0
    except Exception as e:
        print(f"Error reading last serial number: {e}")
        return 0

def initialize_serial_number():
    global current_serial_number, filename  # Add filename to global variables if it's not already there
    
    # Read the configuration file
    config = configparser.ConfigParser()
    config.read('article_config.txt')
    
    # Get the filename from the config
    filename = config['DEFAULT']['filename']
    
    script_dir = os.path.dirname(os.path.abspath(__file__))
    local_log_filepath = os.path.join(script_dir, "Artiklar", f"{filename}_log.xlsx")
    
    current_serial_number = read_last_serial_number_from_log(local_log_filepath)
    print(f"Initialized serial number to: {current_serial_number}")


def split_description(text, max_length):
    if text is None:
        return "", ""
    if len(text) > max_length:
        split_pos = text.rfind(' ', 0, max_length)
        if split_pos == -1:
            split_pos = max_length
        line1 = text[:split_pos]
        line2 = text[split_pos + 1:]
        return line1, line2
    else:
        return text, ""

def print_label(serial_number):
    # Read the configuration file
    config = configparser.ConfigParser()
    config.read('article_config.txt')

    # Get variables from config
    part_number = config['DEFAULT'].get('filename', 'Unknown')
    description = config['DEFAULT'].get('description', '')
    rev = config['DEFAULT'].get('rev', '')

    # Prepare other variables
    week = datetime.now().strftime("%y%W")  # Current year and week number

    # Split description if needed
    description_line_1, description_line_2 = split_description(description, 30)

    # Format serial number to 5 digits
    serial = f"{serial_number:05d}"

    # Adjusted ZPL template with improved positioning
    zpl_template = """
    ^XA
    ^FO00,30^A0N,90,90^FD{PART_NUMBER}^FS  ; Part Number, made larger
    ^FO00,130^A0N,40,40^FD{DESCRIPTION_LINE_1}^FS  ; Description Line 1, adjusted position
    ^FO00,180^A0N,40,40^FD{DESCRIPTION_LINE_2}^FS  ; Description Line 2, adjusted position
    ^FO00,250^A0N,50,50^FDWeek^FS              ; "Week" label, adjusted position
    ^FO240,250^A0N,50,50^FDREV^FS              ; "REV" label, adjusted position
    ^FO400,250^A0N,50,50^FDSerial^FS           ; "Serial" label, adjusted position
    ^FO00,320^A0N,70,70^FD{WEEK}^FS            ; Week number, adjusted position and size
    ^FO240,320^A0N,70,70^FD{REV}^FS            ; Revision, adjusted position and size
    ^FO400,320^A0N,70,70^FD{SERIAL}^FS         ; Serial number, adjusted position and size
    ^XZ
    """

    # Replace placeholders with actual values
    zpl_code = zpl_template.format(
        PART_NUMBER=part_number,
        DESCRIPTION_LINE_1=description_line_1,
        DESCRIPTION_LINE_2=description_line_2,
        WEEK=week,
        REV=rev,
        SERIAL=serial
    )

    # Write the ZPL to a temporary file
    zpl_file_path = "/tmp/label.zpl"
    with open(zpl_file_path, "w") as f:
        f.write(zpl_code)

    # Send the ZPL file to the printer in raw mode
    os.system(f"lp -d GK420d -o raw {zpl_file_path}")

    print(f"Label printed for part number {part_number}, serial number: {serial}")
    print(f"Description: {description}")  # Add this line for debugging


# ... (after global variable declarations)
initialize_serial_number()
# ... (before GUI setup)

# Initialize main window
root = tk.Tk()
root.title("Testen")
root.geometry("1920x1080")
root.attributes('-fullscreen', True)
root.bind("<Escape>", lambda e: root.destroy())  # Allow exiting fullscreen with the Esc key
root.lift()
root.attributes('-topmost', True)
root.after(100, lambda: root.attributes('-topmost', False, '-fullscreen', True))


def custom_messagebox(title, message, box_type="info"):
    custom_box = tk.Toplevel(root)
    custom_box.title(title)
    custom_box.geometry("600x300+600+200")
    custom_box.attributes('-topmost', 'true')
    custom_box.grab_set()
    custom_box.focus_force()

    msg_label = tk.Label(custom_box, text=message, font=("Helvetica", 18), wraplength=550)
    msg_label.pack(pady=40)

    response_var = tk.BooleanVar()  # Use a BooleanVar to store the response

    if box_type == "error" or box_type == "info":
        button_text = "OK"
        button_command = lambda: (response_var.set(True), custom_box.destroy())
        ok_button = tk.Button(custom_box, text=button_text, font=("Helvetica", 18), width=12, height=3, command=button_command)
        ok_button.pack(pady=40)

    elif box_type == "askyesno":
        def on_yes():
            response_var.set(True)
            custom_box.destroy()

        def on_no():
            response_var.set(False)
            custom_box.destroy()

        yes_button = tk.Button(custom_box, text="Yes", font=("Helvetica", 18), width=12, height=3, command=on_yes)
        yes_button.pack(side=tk.LEFT, padx=20)

        no_button = tk.Button(custom_box, text="No", font=("Helvetica", 18), width=12, height=3, command=on_no)
        no_button.pack(side=tk.RIGHT, padx=20)

    custom_box.wait_window()  # Wait for the user to respond
    return response_var.get()  # Return the user’s response as True or False

def custom_info_popup(title, message):
    custom_box = tk.Toplevel(root)
    custom_box.title(title)
    custom_box.geometry("600x300+600+200")
    custom_box.attributes('-topmost', 'true')
    custom_box.grab_set()
    custom_box.focus_force()

    msg_label = tk.Label(custom_box, text=message, font=("Helvetica", 18), wraplength=550)
    msg_label.pack(pady=40)

    ok_button = tk.Button(custom_box, text="OK", font=("Helvetica", 18), width=12, height=3, command=custom_box.destroy)
    ok_button.pack(pady=40)

    custom_box.update()
    custom_box.wait_window()




print(f"{datetime.now()}: Window opened")



# Initialize I2C bus and MCP23017
i2c = busio.I2C(board.SCL, board.SDA)
mcp1 = MCP23017(i2c, address=0x20)
mcp2 = MCP23017(i2c, address=0x22)
relay_mcp1 = MCP23017(i2c, address=0x21)
relay_mcp2 = MCP23017(i2c, address=0x23)


def enable_probing():
    global expecting_probe
    expecting_probe = True
    print("Probing is now enabled.")

expecting_probe = False


# Configure relay pins
relay_mappings = {
    '1: A': (relay_mcp1, 0),   # MCP 0x21 pin A0
    '2: B': (relay_mcp1, 1),   # MCP 0x21 pin A1
    '3: C': (relay_mcp1, 2),   # MCP 0x21 pin A2
    '4: D': (relay_mcp1, 3),   # MCP 0x21 pin A3
    '5: E': (relay_mcp1, 4),   # MCP 0x21 pin A4
    '6: F': (relay_mcp1, 5),   # MCP 0x21 pin A5
    '7: G': (relay_mcp1, 6),   # MCP 0x21 pin A6
    '8: H': (relay_mcp1, 7),   # MCP 0x21 pin A7
    '9: J': (relay_mcp1, 8),   # MCP 0x21 pin B0
    '10: K': (relay_mcp1, 9),  # MCP 0x21 pin B1
    '11: L': (relay_mcp1, 10), # MCP 0x21 pin B2
    '12: M': (relay_mcp1, 11), # MCP 0x21 pin B3
    '13: N': (relay_mcp1, 12), # MCP 0x21 pin B4
    '14: P': (relay_mcp1, 13), # MCP 0x21 pin B5
    '15: R': (relay_mcp1, 14), # MCP 0x21 pin B6
    '16: S': (relay_mcp1, 15), # MCP 0x21 pin B7
    '17: T': (relay_mcp2, 0),  # MCP 0x23 pin A0
    '18: U': (relay_mcp2, 1),  # MCP 0x23 pin A1
    '19: V': (relay_mcp2, 2),  # MCP 0x23 pin A2
    '20: W': (relay_mcp2, 3),  # MCP 0x23 pin A3
    '21: X': (relay_mcp2, 4),  # MCP 0x23 pin A4
    '22: Y': (relay_mcp2, 5),  # MCP 0x23 pin A5
    '23: Z': (relay_mcp2, 6),  # MCP 0x23 pin A6
    '24: a': (relay_mcp2, 7),  # MCP 0x23 pin A7
    '25: b': (relay_mcp2, 8),  # MCP 0x23 pin B0
    '26: c': (relay_mcp2, 9),  # MCP 0x23 pin B1
    '27: d': (relay_mcp2, 10), # MCP 0x23 pin B2
    '28: e': (relay_mcp2, 11), # MCP 0x23 pin B3
    '29: f': (relay_mcp2, 12), # MCP 0x23 pin B4
    '30: g': (relay_mcp2, 13), # MCP 0x23 pin B5
    '31: h': (relay_mcp2, 14), # MCP 0x23 pin B6
    '32: j': (relay_mcp2, 15)  # MCP 0x23 pin B7
}


def activate_relay(pin_label):
    if pin_label in relay_mappings:
        mcp, pin = relay_mappings[pin_label]
        relay_pin = mcp.get_pin(pin)
        relay_pin.value = True  # Activate relay by setting it high
        print(f"Activated relay for {pin_label}")

def deactivate_relay(pin_label):
    if pin_label in relay_mappings:
        mcp, pin = relay_mappings[pin_label]
        relay_pin = mcp.get_pin(pin)
        relay_pin.value = False  # Deactivate relay by setting it low
        print(f"Deactivated relay for {pin_label}")

# Ensure to deactivate all relays on startup
for pin_label in relay_mappings.keys():
    deactivate_relay(pin_label)

# List of MCP23017 pins to test (both chips)
mcp_pins = [(mcp1, i) for i in range(16)] + [(mcp2, i) for i in range(16)]
for mcp, pin in mcp_pins:
    mcp_pin = mcp.get_pin(pin)
    mcp_pin.direction = Direction.INPUT
    mcp_pin.pull = Pull.UP

# Initialize relay pins
for mcp, pin in relay_mappings.values():
    relay_pin = mcp.get_pin(pin)
    relay_pin.direction = Direction.OUTPUT
    relay_pin.value = False  # Assuming relay is off when low




# Read the configuration file
config = configparser.ConfigParser()
config.read('article_config.txt')

filename = config['DEFAULT']['filename']
description = config['DEFAULT']['description']
pins = config['DEFAULT']['pins'].split(',')
if pins[0].startswith("pins="):
    pins[0] = pins[0][5:]
takeouts = int(config['DEFAULT']['takeouts'])
stops = takeouts - 1  # Adjusting the number of stops to account for manual winding to the first takeout
spacing = float(config['DEFAULT']['spacing']) * 1000  # Convert to mm
length = float(config['DEFAULT']['length']) * 1000  # Convert to mm
inner_diameter = float(config['DEFAULT']['inner_diameter'])
width = float(config['DEFAULT']['width'])

# Display the parameters
print(f"{datetime.now()}: Config read successfully.")
print(f"Filename: {filename}")
print(f"Description: {description}")
print(f"Pins: {pins}")
print(f"Takeouts: {takeouts}")
print(f"Spacing: {spacing} mm")
print(f"Total Length: {length} mm")
print(f"Inner Diameter: {inner_diameter} mm")
print(f"Drum Width: {width} mm")

# Get the directory where the script is located
script_dir = os.path.dirname(os.path.abspath(__file__))

# Construct the full path to the image file inside the PI folder
logo_path = os.path.join(script_dir, "logo.png")

# Variable Information
amount_of_cycles_done = 0
elapsed_time_current_cycle = 0
elapsed_time_previous_cycle = 0
total_elapsed_time = 0
downtime = 0
skipped_tests = 0
current_pin_index = 0
is_running = False


#Custom fonts
header_font = font.Font(family="Helvetica", size=24, weight="bold")
center_font = font.Font(family="Helvetica", size=124, weight="bold")
body_font = font.Font(family="Helvetica", size=16)

def format_time(seconds):
    minutes = seconds // 60
    seconds = seconds % 60
    return f"{minutes}:{seconds:02}"

def seconds_to_hms(seconds):
    # Converts seconds to HH:MM:SS format
    return str(timedelta(seconds=seconds))

def str_to_timedelta(time_str):
    if isinstance(time_str, timedelta):
        return time_str
    try:
        time_obj = datetime.strptime(time_str, "%H:%M:%S")
        return timedelta(hours=time_obj.hour, minutes=time_obj.minute, seconds=time_obj.second)
    except ValueError:
        return timedelta(seconds=0)

def mcp_pin_to_gui_pin(mcp, pin):
    mapping = {
        (0x20, 0): '1: A',  # MCP 0x20 pin A0 to GUI pin 1: A
        (0x20, 1): '2: B',  # MCP 0x20 pin A1 to GUI pin 2: B
        (0x20, 2): '3: C',  # MCP 0x20 pin A2 to GUI pin 3: C
        (0x20, 3): '4: D',  # MCP 0x20 pin A3 to GUI pin 4: D
        (0x20, 4): '5: E',  # MCP 0x20 pin A4 to GUI pin 5: E
        (0x20, 5): '6: F',  # MCP 0x20 pin A5 to GUI pin 6: F
        (0x20, 6): '7: G',  # MCP 0x20 pin A6 to GUI pin 7: G
        (0x20, 7): '8: H',  # MCP 0x20 pin A7 to GUI pin 8: H
        (0x20, 8): '16: S', # MCP 0x20 pin B0 to GUI pin 16: S
        (0x20, 9): '15: R', # MCP 0x20 pin B1 to GUI pin 15: R
        (0x20, 10): '14: P',# MCP 0x20 pin B2 to GUI pin 14: P
        (0x20, 11): '13: N',# MCP 0x20 pin B3 to GUI pin 13: N
        (0x20, 12): '12: M',# MCP 0x20 pin B4 to GUI pin 12: M
        (0x20, 13): '11: L',# MCP 0x20 pin B5 to GUI pin 11: L
        (0x20, 14): '10: K',# MCP 0x20 pin B6 to GUI pin 10: K
        (0x20, 15): '9: J', # MCP 0x20 pin B7 to GUI pin 9: J
        (0x22, 0): '17: T', # MCP 0x22 pin A0 to GUI pin 17: T
        (0x22, 1): '18: U', # MCP 0x22 pin A1 to GUI pin 18: U
        (0x22, 2): '19: V', # MCP 0x22 pin A2 to GUI pin 19: V
        (0x22, 3): '20: W', # MCP 0x22 pin A3 to GUI pin 20: W
        (0x22, 4): '21: X', # MCP 0x22 pin A4 to GUI pin 21: X
        (0x22, 5): '22: Y', # MCP 0x22 pin A5 to GUI pin 22: Y
        (0x22, 6): '23: Z', # MCP 0x22 pin A6 to GUI pin 23: Z
        (0x22, 7): '24: a', # MCP 0x22 pin A7 to GUI pin 24: a
        (0x22, 8): '25: b', # MCP 0x22 pin B0 to GUI pin 25: b
        (0x22, 9): '26: c', # MCP 0x22 pin B1 to GUI pin 26: c
        (0x22, 10): '27: d',# MCP 0x22 pin B2 to GUI pin 27: d
        (0x22, 11): '28: e',# MCP 0x22 pin B3 to GUI pin 28: e
        (0x22, 12): '29: f',# MCP 0x22 pin B4 to GUI pin 29: f
        (0x22, 13): '30: g',# MCP 0x22 pin B5 to GUI pin 30: g
        (0x22, 14): '31: h',# MCP 0x22 pin B6 to GUI pin 31: h
        (0x22, 15): '32: j',# MCP 0x22 pin B7 to GUI pin 32: j
    }
    return mapping.get((mcp, pin), None)




def read_mcp_probes():
    if not is_running or not expecting_probe:
        return None, None, False
    for mcp_chip, pin in mcp_pins:
        mcp_address = mcp_chip._device.device_address  # Updated to use _device
        mcp_pin = mcp_chip.get_pin(pin)
        if not mcp_pin.value:  # LOW indicates probed
            print(f"Detected LOW signal: MCP={hex(mcp_address)}, PIN={pin}")  # Updated logging
            return mcp_address, pin, True
    return None, None, False


color_mapping = {
    "1: A": "White",
    "2: B": "#9c451e",
    "3: C": "#32CD32",
    "4: D": "Yellow",
    "5: E": "#b1afb3",
    "6: F": "Pink",
    "7: G": "#0A60C5",
    "8: H": "Red",
    "9: J": "#353436",
    "10: K": "#c934eb",
    "11: L": ("Grey", "Pink"),
    "12: M": ("Red", "Blue"),
    "13: N": ("White", "Green"),
    "14: P": ("Brown", "Green"),
    "15: R": ("White", "Yellow"),
    "16: S": ("Yellow", "Brown"),
    "17: T": ("White", "Grey"),
    "18: U": ("Grey", "Brown"),
    "19: V": ("White", "Pink"),
    "20: W": ("Pink", "Brown"),
    "21: X": ("White", "Blue"),
    "22: Y": ("Brown", "Blue"),
    "23: Z": ("White", "Red"),
    "24: a": ("Brown", "Red"),
    "25: b": ("White", "Black"),
    "26: c": ("Brown", "Black"),
    "27: d": ("Grey", "Green"),
    "28: e": ("Yellow", "Grey"),
    "29: f": ("Pink", "Green"),
    "30: g": ("Yellow", "Pink"),
    "31: h": ("Green", "Blue"),
    "32: j": ("Yellow", "Blue"),
}
def color_to_rgb(color):
    # Complete color map including all possible color names used in color_mapping
    color_map = {
        "White": (255, 255, 255),
        "Brown": (156, 69, 30),
        "Green": (85, 235, 52),
        "Yellow": (255, 255, 0),
        "Grey": (128, 128, 128),
        "Pink": (255, 192, 203),
        "Blue": (0, 0, 255),
        "Red": (255, 0, 0),
        "Black": (53, 52, 54),
        "Violet": (201, 52, 235)
    }
    return color_map.get(color, (0, 0, 0))  # Default to black if color not found

    

def set_dual_color(label, color1, color2=None, pin_text="", width=600, height=500):
    # Convert the color names to RGB tuples
    color1_rgb = color_to_rgb(color1)
    color2_rgb = color_to_rgb(color2) if color2 else color1_rgb

    # Log the colors being used for debugging
    print(f"Setting dual color for center label:")
    print(f"  Primary Color: {color1} -> RGB: {color1_rgb}")
    print(f"  Secondary Color: {color2} -> RGB: {color2_rgb}")
    print(f"Pin Text being set: {pin_text}")

    # Create a new image for the gradient
    width, height = 600, 500  # Adjust dimensions as needed
    gradient_image = Image.new("RGB", (width, height))

    for y in range(height):
        for x in range(width):
            if x < width * 0.66:
                gradient_image.putpixel((x, y), color1_rgb)
            else:
                gradient_image.putpixel((x, y), color2_rgb)

    # Convert the image to a PhotoImage and set it as the label's background
    gradient_photo = ImageTk.PhotoImage(gradient_image)

    # Set the image and text on the label
    label.config(image=gradient_photo, text=pin_text, compound='center', width=width, height=height)
    label.image = gradient_photo  # Keep a reference to avoid garbage collection

    # Force the UI to update immediately
    label.update_idletasks()







# Modify the on_pin_probe and complete_probe functions
allow_motor_run = True

# Modify the on_pin_probe function
def on_pin_probe(gui_pin_label):
    global current_pin_index, expecting_probe, success_sound, reject_sound, current_segment, allow_motor_run

    print(f"on_pin_probe called with gui_pin_label={gui_pin_label}, expecting_probe={expecting_probe}")

    if not expecting_probe:
        print("Probe not expected, ignoring...")
        return

    expected_pin_label = left_panel_labels[current_pin_index].cget("text")

    if gui_pin_label == expected_pin_label:
        print("Probe matches expected pin")
        expecting_probe = False  # Disable further probing until the next pin

        if success_sound:
            success_sound.play()

        if (current_pin_index % (len(left_panel_labels) // takeouts) == 0) or (current_pin_index == len(left_panel_labels)):
            allow_motor_run = True
        else:
            allow_motor_run = False

        left_panel_labels[current_pin_index].config(bg="#32CD32")
        deactivate_relay(expected_pin_label)  # Deactivate previous relay
        current_pin_index += 1

        if current_pin_index < len(left_panel_labels):
            next_pin_label = left_panel_labels[current_pin_index].cget("text")
            # Update the central label color and add debugging output
            if isinstance(color_mapping[next_pin_label], tuple):
                print(f"Next pin uses dual color: {color_mapping[next_pin_label]}")
                set_dual_color(current_wire_label, *color_mapping[next_pin_label], pin_text=next_pin_label)
            else:
                print(f"Next pin uses single color: {color_mapping[next_pin_label]}")
                current_wire_label.config(text=next_pin_label, bg=color_mapping[next_pin_label])

            left_panel_labels[current_pin_index].config(bg="yellow")
        else:
            print("All pins probed successfully.")
            check_all_probed()

        # Check if the segment is complete and update motor button
        if current_pin_index % (len(left_panel_labels) // takeouts) == 0:
            print("Probing for the current segment is complete. Updating motor button.")
            allow_motor_run = True
            update_motor_button()  # Update the motor button state to reflect readiness
        else:
            # If not at the end of a segment, allow continuation of probing
            allow_motor_run = False

    else:
        print(f"Pin mismatch: expected {expected_pin_label}, but got {gui_pin_label}")
        allow_motor_run = False
        update_motor_button()
        if reject_sound:
            reject_sound.play()

    # After probing is complete, update the motor button state
    if current_pin_index % (len(left_panel_labels) // takeouts) == 0:
        print("Probing for the current segment is complete. Updating motor button.")
        update_motor_button()  # Update the motor button state to reflect readiness



def activate_relay_and_wait(pin_label):
    print(f"Activating relay for {pin_label}")
    activate_relay(pin_label)
    root.after(2000, lambda: start_probing(pin_label))  # Wait 2 seconds before probing


# Function to calculate rotations for each segment
def calculate_rotations():
    global rotation_list  # Ensure we're modifying the global rotation_list
    # Calculate the length wound before the motor-controlled segments start
    length_wound = (length - spacing * (takeouts - 1)) / 2
    layers_wound = length_wound / (width * CABLE_DIAMETER)
    diameter_increase = 2 * CABLE_DIAMETER * layers_wound
    current_diameter = inner_diameter + diameter_increase
    
    segment_length = spacing  # Length of cable per segment
    rotation_list = []  # Reset the list

    # Loop for the calculated number of stops
    for i in range(takeouts - 1):
        # Calculate the current circumference
        current_circumference = math.pi * current_diameter

        # Calculate the number of rotations for this segment
        rotations = segment_length / current_circumference
        rotation_list.append(rotations)

        # Print details for verification
        print(f"Segment {i + 1}:")
        print(f"  Current Diameter: {current_diameter:.2f} mm")
        print(f"  Current Circumference: {current_circumference:.2f} mm")
        print(f"  Segment Length: {segment_length:.2f} mm")
        print(f"  Rotations Needed: {rotations:.2f}")

        # Increase the diameter for the next layer
        layers = math.floor(width / CABLE_DIAMETER)
        current_diameter += 2 * CABLE_DIAMETER * layers
    
    # Debug print to verify the rotation list
    print(f"Debug: rotation_list = {rotation_list}")

    return rotation_list

# Ensure rotation_list is populated
rotation_list = calculate_rotations()

# Debug log to verify rotation_list is populated
print(f"Debug: rotation_list = {rotation_list}")

def update_motor_button():
    global motor_button, current_segment, rotation_list, allow_motor_run

    if allow_motor_run and current_segment < len(rotation_list):
        if current_segment == len(rotation_list) - 1:
            # Final segment
            motor_button.config(
                text=f"Motor\nFinal segment\n{rotation_list[current_segment]:.2f} rotations",
                bg="green", state=tk.NORMAL
            )
        else:
            motor_button.config(
                text=f"Motor\n{current_segment + 1}/{len(rotation_list)} segment\n{rotation_list[current_segment]:.2f} rotations",
                bg="green", state=tk.NORMAL
            )
    else:
        motor_button.config(text="Motor", bg="gray", state=tk.DISABLED)



def run_motor():
    global current_segment, allow_motor_run
    if current_segment < len(rotation_list):
        rotations = rotation_list[current_segment]
        print(f"Running Motor for Segment {current_segment + 1}:")
        print(f"  Rotations to perform: {rotations:.2f}")

        # Publish the number of rotations to the MQTT broker
        client.publish("motor/control", str(rotations))

        current_segment += 1  # Move to the next segment
        allow_motor_run = False
        if current_segment < len(rotation_list):
            enable_probing()
        else:
            # Test complete
            print("All segments completed.")
            # Add any end-of-test logic here
        update_motor_button()  # Update the button after running the motor

        # Turn the button gray after use
        motor_button.config(text="Motor", bg="gray")
        motor_button.config(state=tk.DISABLED)  # Disable the button

        # Enable probing for the next pin
        enable_probing()
    else:
        print("No more segments left to run.")






def monitor_pins():
    print("Starting pin monitoring...")  # Initial log to confirm thread start
    while True:
        mcp, pin, probe_value = read_mcp_probes()
        if probe_value:
            gui_pin_label = mcp_pin_to_gui_pin(mcp, pin)
            if gui_pin_label is not None:
                print(f"Probing detected: MCP={hex(mcp)}, PIN={pin}, GUI_PIN={gui_pin_label}")  # Detailed log
                root.after(0, lambda: on_pin_probe(gui_pin_label))
            else:
                print(f"No mapping found for MCP={hex(mcp)}, PIN={pin}")  # Log for missing mapping
        time.sleep(0.1)



# Start monitoring pins in a separate thread
thread = threading.Thread(target=monitor_pins)
thread.daemon = True
thread.start()

def calculate_average(ws):
    total_time = timedelta()
    total_count = 0
    for row in range(6, ws.max_row + 1):
        total_time += str_to_timedelta(ws[f'E{row}'].value)
        total_count += ws[f'C{row}'].value

    if total_count > 0:
        average_time = total_time / total_count
    else:
        average_time = timedelta()
    return average_time

def confirm_last_probe():
    global current_pin_index
    if current_pin_index == len(pins) - 1:
        print(f"All pins probed, checking all probed status in 500ms")
        root.after(500, check_all_probed)

def check_all_probed():
    all_probed = all(label.cget("bg") == "#32CD32" for label in left_panel_labels)
    print(f"Check all probed: {all_probed}")
    for i, label in enumerate(left_panel_labels):
        print(f"Pin {i+1} background color: {label.cget('bg')}")
    if all_probed:
        complete_cycle()
    else:
        confirm_complete_cycle()



def complete_probe():
    global current_pin_index
    if current_pin_index < len(pins) - 1:
        left_panel_labels[current_pin_index].config(bg="#32CD32")
        print(f"Pin {pins[current_pin_index]} probed successfully, setting to green")
        current_pin_index += 1

        next_pin_label = pins[current_pin_index]
        # Update the central label color
        if isinstance(color_mapping[next_pin_label], tuple):
            set_dual_color(current_wire_label, *color_mapping[next_pin_label])
        else:
            current_wire_label.config(text=next_pin_label, bg=color_mapping[next_pin_label])

        left_panel_labels[current_pin_index].config(bg="yellow")
        print(f"Next pin to probe: {next_pin_label}")
    else:
        left_panel_labels[current_pin_index].config(bg="#32CD32")
        print(f"Last pin probed: {pins[current_pin_index]}. Confirming last probe in 500ms")
        root.after(500, confirm_last_probe)

def complete_cycle():
    global amount_of_cycles_done, elapsed_time_current_cycle, elapsed_time_previous_cycle, total_elapsed_time, current_pin_index, is_running, skipped_tests, current_segment, allow_motor_run, current_serial_number

    amount_of_cycles_done += 1
    current_serial_number += 1
    elapsed_time_previous_cycle = elapsed_time_current_cycle
    total_elapsed_time += elapsed_time_previous_cycle
    elapsed_time_current_cycle = 0
    completed_label.config(text=f"Färdiga: {amount_of_cycles_done}st")
    time_info_label.config(text=f"Tid\nNu: {format_time(elapsed_time_current_cycle)}\nFörra: {format_time(elapsed_time_previous_cycle)}\nTotal: {format_time(total_elapsed_time)}\nStälltid: {format_time(downtime)}")

    # Reset the segment index and allow motor run at the start of a new cycle
    current_segment = 0
    allow_motor_run = True
    
    # Reset pins for the next cycle
    current_pin_index = 0
    for label in left_panel_labels:
        label.config(bg="light gray")  # Reset to default background color

    # Reset the central label to its original state
    current_wire_label.config(text="Starta", bg="#32CD32", font=center_font)
    current_wire_label.config(width=6, height=3, image='')  # Clear any image and reset size

    is_running = False
    print(f"Cycle completed successfully. Pins reset for next cycle.")

    # Print label with serial number
    print_label(current_serial_number)

    # The cycle data logging is already handled in your existing batch sheet update


def confirm_complete_cycle():
    print("Opening confirmation window for incomplete cycle.")
    confirm_window = tk.Toplevel(root)
    confirm_window.title("Bekräfta")
    confirm_window.geometry("1920x1080")
    confirm_window.attributes('-fullscreen', True)

    message_label = tk.Label(confirm_window, text="Du försöker färdigställa utan att alla punkter är testade, är du säker att du vill fortsätta?\nBekräfta genom att skriva OK", font=body_font)
    message_label.pack(pady=10)

    entry = tk.Entry(confirm_window, font=body_font)
    entry.pack(pady=10)

    def on_confirm():
        if entry.get().strip().upper() == "OK":
            confirm_window.destroy()
            complete_cycle()
        else:
            tk.messagebox.showerror("Fel", "Felaktig bekräftelse")

    confirm_button = tk.Button(confirm_window, text="Bekräfta", font=body_font, command=on_confirm)
    confirm_button.pack(pady=10)

    # On-screen keyboard
    keyboard_frame = tk.Frame(confirm_window, bg="blue")
    keyboard_frame.pack(pady=10,fill=tk.BOTH, expand=True)

    def insert_char(char):
        entry.insert(tk.END, char)

    keys = [
        'QWERTYUIOP',
        'ASDFGHJKL',
        'ZXCVBNM'
    ]

    key_padx = 5  # Padding for keys
    key_pady = 5

    for i, key_row in enumerate(keys):
        row_frame = tk.Frame(keyboard_frame, bg="blue")
        row_frame.grid(row=i, column=0, columnspan=3, pady=key_pady, padx=key_padx)
        for char in key_row:
            button = tk.Button(row_frame, text=char, font=body_font, width=6, height=4, command=lambda c=char: insert_char(c))
            button.pack(side=tk.LEFT, padx=key_padx, pady=key_pady)

    space_button = tk.Button(keyboard_frame, text="SPACE", font=body_font, width=20, height=2, command=lambda: insert_char(" "))
    space_button.grid(row=3, column=0, columnspan=3, pady=key_pady)

    # Arrange numpad keys in the keyboard window
    numpad_keys = [
        '789',
        '456',
        '123',
        '0+-'
    ]

    numpad_frame = tk.Frame(keyboard_frame, bg="blue")
    numpad_frame.grid(row=0, column=4, rowspan=4, padx=key_padx, pady=key_pady)

    for i, key_row in enumerate(numpad_keys):
        row_frame = tk.Frame(numpad_frame, bg="blue")
        row_frame.pack(pady=key_pady)
        for char in key_row:
            button = tk.Button(row_frame, text=char, font=body_font, width=8, height=5, command=lambda c=char: insert_char(c))
            button.pack(side=tk.LEFT, padx=key_padx, pady=key_pady)

    close_button = tk.Button(keyboard_frame, text="X", font=body_font, width=10, height=5, command=confirm_window.destroy, bg="red")
    close_button.grid(row=5, column=4, columnspan=3, pady=key_pady)





def update_timer():
    global elapsed_time_current_cycle, downtime
    if is_running:
        elapsed_time_current_cycle += 1
    else:
        downtime += 1
    time_info_label.config(text=f"Tid\nNu: {format_time(elapsed_time_current_cycle)}\nFörra: {format_time(elapsed_time_previous_cycle)}\nTotal: {format_time(total_elapsed_time + elapsed_time_current_cycle)}\nStälltid: {format_time(downtime)}")
    root.after(1000, update_timer)  # Update every second (1000 milliseconds)

def reset_test():
    response = custom_messagebox("Reset", "Är du säker att du vill reseta?")
    if response:
        global current_pin_index, elapsed_time_current_cycle, total_elapsed_time, downtime, is_running, expecting_probe, current_segment, allow_motor_run

        # Reset the segment index and allow motor run at the start of a new cycle
        current_segment = 0
        allow_motor_run = True

        total_elapsed_time += elapsed_time_current_cycle
        elapsed_time_current_cycle = 0
        current_pin_index = 0
        downtime = 0
        is_running = False
        expecting_probe = False
        for label in left_panel_labels:
            label.config(bg="light gray")
        left_panel_labels[current_pin_index].config(bg="yellow")

        # Reset the central label size and appearance
        current_wire_label.config(text="Starta", bg="#32CD32", font=center_font)
        current_wire_label.config(width=6, height=3, image='')  # Clear the image and reset the size

        time_info_label.config(text=f"Tid\nNu: {format_time(elapsed_time_current_cycle)}\nFörra: {format_time(elapsed_time_previous_cycle)}\nTotal: {format_time(total_elapsed_time)}\nStälltid: {format_time(downtime)}")

        # Reset all MCP23017 pins
        for mcp, pin in mcp_pins:
            mcp_pin = mcp.get_pin(pin)
            mcp_pin.direction = Direction.INPUT
            mcp_pin.pull = Pull.UP

        print("Reset complete and MCP23017 pins reset")  # Add logging for debugging

        # Update the motor button
        update_motor_button()

        print("Reset complete. All pins and labels should be back to the initial state.")




# Global variable to track relay state
relay_active = False



def toggle_timer():
    global is_running
    is_running = not is_running
    if is_running:
        current_wire_label.config(bg="yellow", text=pins[current_pin_index])
        print("System unpaused. Ready to probe.")
        # No automatic relay activation on start
        root.after(1000, lambda: start_probing(pins[current_pin_index]))  # Allow probing after unpausing
    else:
        print(f"System paused at pin {pins[current_pin_index]}")
        current_wire_label.config(bg="orange", text="Pausad")

def manual_relay_control():
    global relay_active, expecting_probe, allow_motor_run
    
    if relay_active:
        # If the relay is currently active, deactivate it and then activate probing
        deactivate_relay(pins[current_pin_index])
        relay_active = False
        print(f"Relay manually deactivated for {pins[current_pin_index]}")
        
        # Activate probing after the relay is deactivated
        expecting_probe = True
        print("Probing is now active.")
        
    else:
        # If the relay is not active, activate it and deactivate probing
        activate_relay(pins[current_pin_index])
        relay_active = True
        print(f"Relay manually activated for {pins[current_pin_index]}")
        
        # Deactivate probing while the relay is active
        expecting_probe = False
        print("Probing is now deactivated.")
       # Allow the motor to run only if the segment is completed
    if allow_motor_run:
        update_motor_button()

def start_probing(pin_label):
    global expecting_probe
    expecting_probe = True
    print(f"Probing started for {pin_label}")



def deactivate_relay_and_wait_for_probe(pin_label):
    deactivate_relay(pin_label)  # Deactivate relay after 2 seconds
    print(f"Relay deactivated for {pin_label}, now waiting for probe.")

def on_pin_click(idx):
    global current_pin_index
    if idx > current_pin_index:
        # Check if the jump skips any pins in the current cycle
        for i in range(current_pin_index + 1, idx + 1):
            if left_panel_labels[i].cget("bg") != "#32CD32":  # Not green
                response = custom_info_popup("Hoppa över", "Du hoppar över flera punkter, är du säker att du vill fortsätta?", "askyesno")
                if not response:
                    return
                break
    # Preserve the green status if the pin was already tested
    current_color = left_panel_labels[current_pin_index].cget("bg")
    if current_color != "#32CD32":
        left_panel_labels[current_pin_index].config(bg="light gray")
    
    current_pin_index = idx
    left_panel_labels[current_pin_index].config(bg="yellow")
    current_wire_label.config(text=pins[current_pin_index], bg="yellow")
    if not is_running:
        current_wire_label.config(bg="#32CD32", text="Starta")

def time_string_to_seconds(time_str):
    if isinstance(time_str, int):
        return time_str
    time_obj = datetime.strptime(time_str, "%H:%M:%S")
    return timedelta(hours=time_obj.hour, minutes=time_obj.minute, seconds=time_obj.second).total_seconds()

def calculate_totals(data):
    total_cycles = sum(data['Antal'])
    total_time = sum([str_to_timedelta(time_str).total_seconds() for time_str in data['Total Cykeltid (HH:MM:SS)']])
    return total_cycles, total_time

def calculate_average_time(total_time, total_cycles):
    if total_cycles > 0:
        average_time = timedelta(seconds=total_time / total_cycles)
    else:
        average_time = timedelta(seconds=0)
    return average_time





def save_log(filename, data):
    # Convert the data dictionary to a DataFrame
    df = pd.DataFrame(data)

    # Create a new workbook and add a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # Add the image to the worksheet
    img = Image(logo_path)
    img.anchor = 'A1'
    ws.add_image(img)


    # Add column headers for the log data
    column_headers = ["Batchdatum", "Antal", "Antal skippad test", "Total Cykeltid (HH:MM:SS)",
                      "Total Ställtid (HH:MM:SS)", "Total Stycktid (HH:MM:SS)", "Cykeltid (HH:MM:SS)",
                      "Stycktid (HH:MM:SS)", "Styck Ställtid (HH:MM:SS)"]
    
    for col_num, header in enumerate(column_headers, 1):
        col_letter = openpyxl.utils.get_column_letter(col_num)
        ws[f'{col_letter}6'] = header
        ws[f'{col_letter}6'].alignment = Alignment(horizontal="center")
        ws[f'{col_letter}6'].font = openpyxl.styles.Font(bold=True)

    # Add the data rows to the worksheet starting from row 7
    for row_num, row_data in df.iterrows():
        for col_num, (col_name, value) in enumerate(row_data.items(), 1):
            col_letter = openpyxl.utils.get_column_letter(col_num)
            ws[f'{col_letter}{row_num + 7}'] = value

    # Save the workbook
    wb.save(filename)


def create_new_log_file(filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Main"

    # Extract the relevant part of the filename
    base_filename = os.path.basename(filename)
    article_number = base_filename.split('_')[0]

    # Add the image and set its position and size
    img = OpenpyxlImage(logo_path)
    img.anchor = 'A1'
    img.width = 340  # Set the width to fit into A1:B3
    img.height = 80  # Set the height to fit into A1:B3
    ws.add_image(img)

    # Add headers
    headers = ["Artikelnummer", "AVG. Stycktid", "Senaste Stycktid"]
    values = [article_number, "00:00:00", "00:00:00"]
    for col_num, (header, value) in enumerate(zip(headers, values), 2):
        ws.cell(row=1, column=col_num * 2).value = header
        ws.cell(row=2, column=col_num * 2).value = value
        ws.cell(row=1, column=col_num * 2).font = Font(bold=True)
        ws.cell(row=2, column=col_num * 2).alignment = Alignment(horizontal="center")

    # Add headers for the log data
    column_headers = ["Batchdatum", "Antal", "Antal skippad test", "Total Cykeltid (HH:MM:SS)",
                      "Total Ställtid (HH:MM:SS)", "Total Stycktid (HH:MM:SS)", "Cykeltid (HH:MM:SS)",
                      "Stycktid (HH:MM:SS)", "Styck Ställtid (HH:MM:SS)"]
    
    for col_num, header in enumerate(column_headers, 2):
        ws.cell(row=5, column=col_num).value = header
        ws.cell(row=5, column=col_num).alignment = Alignment(horizontal="center")
        ws.cell(row=5, column=col_num).font = Font(bold=True)

    # Set column widths
    for col in range(1, 12):
        ws.column_dimensions[get_column_letter(col)].width = 25

    ws.column_dimensions['A'].width = 5   # Adjust as needed
    ws.column_dimensions['B'].width = 20  # Adjust as needed
    ws.column_dimensions['C'].width = 25  # Adjust as needed
    ws.column_dimensions['D'].width = 25  # Adjust as needed
    ws.column_dimensions['E'].width = 25  # Adjust as needed
    ws.column_dimensions['F'].width = 25  # Adjust as needed
    ws.column_dimensions['G'].width = 25  # Adjust as needed
    ws.column_dimensions['H'].width = 25  # Adjust as needed
    ws.column_dimensions['I'].width = 25  # Adjust as needed
    ws.column_dimensions['J'].width = 25  # Adjust as needed
    ws.column_dimensions['K'].width = 25  # Adjust as needed

    # Add headers for total calculations
    ws['G1'] = "Total Tid"
    ws['G3'] = "Total Antal"
    ws['G1'].font = Font(bold=True)
    ws['G3'].font = Font(bold=True)

    # Initialize total values
    ws['G2'] = "00:00:00"
    ws['G4'] = 0

    wb.save(filename)
    print(f"Created new log file: {filename}")

def update_log(filename, data, is_cycle_data=False):
    try:
        if not os.path.exists(filename):
            create_new_log_file(filename)
        
        wb = openpyxl.load_workbook(filename)
        ws_main = wb['Main']

        if is_cycle_data:
            # Handle cycle data in a separate sheet
            batch_date = data["Batchdatum"]
            sheet_name = f"Batch_{batch_date.replace(':', '-')}"
            if sheet_name not in wb.sheetnames:
                ws_batch = wb.create_sheet(title=sheet_name)
                headers = ["Tillverkad", "Antal pins", "Fullt testad", "Serienummer"]
                for col, header in enumerate(headers, 1):
                    ws_batch.cell(row=1, column=col, value=header).font = Font(bold=True)

                ws_batch.column_dimensions['A'].width = 20  # Tillverkad
                ws_batch.column_dimensions['B'].width = 15  # Antal pins
                ws_batch.column_dimensions['C'].width = 15  # Fullt testad
                ws_batch.column_dimensions['D'].width = 15  # Serienummer
            else:
                ws_batch = wb[sheet_name]

            ws_batch.column_dimensions['A'].width = 20  # Tillverkad
            ws_batch.column_dimensions['B'].width = 15  # Antal pins
            ws_batch.column_dimensions['C'].width = 15  # Fullt testad
            ws_batch.column_dimensions['D'].width = 15  # Serienummer
            next_row = ws_batch.max_row + 1
            total_completed = ws_main['G4'].value if ws_main['G4'].value else 0
            serial_number = total_completed + data["Serienummer"]

            ws_batch.cell(row=next_row, column=1, value=data["Batchdatum"])
            ws_batch.cell(row=next_row, column=2, value=len(pins))
            ws_batch.cell(row=next_row, column=3, value="Ja" if data["Antal skippad test"] == 0 else "Nej")
            ws_batch.cell(row=next_row, column=4, value=serial_number)
        else:
            # Update main sheet
            next_row = ws_main.max_row + 1
            if isinstance(data, dict):
                for idx, (key, value) in enumerate(data.items(), start=1):
                    if idx == 1:  # Batchdatum
                        sheet_name = f"Batch_{value.replace(':', '-')}"
                        cell = ws_main.cell(row=next_row, column=idx + 1, value=value)
                        cell.hyperlink = f"#'{sheet_name}'!A1"
                        cell.style = "Hyperlink"
                    else:
                        ws_main.cell(row=next_row, column=idx + 1, value=value)
            elif isinstance(data, list):
                for idx, value in enumerate(data, start=2):
                    ws_main.cell(row=next_row, column=idx, value=value)

            # Update totals and averages
            current_total_time = str_to_timedelta(ws_main['G2'].value) if ws_main['G2'].value else timedelta()
            current_total_count = ws_main['G4'].value if ws_main['G4'].value else 0

            new_time = str_to_timedelta(ws_main[f'E{next_row}'].value)
            new_count = ws_main[f'C{next_row}'].value

            ws_main['G2'] = str(current_total_time + new_time)
            ws_main['G4'] = current_total_count + new_count

            # Update Senaste Stycktid
            ws_main['H2'] = ws_main.cell(row=next_row, column=8).value

            # Calculate average
            total_time = sum([str_to_timedelta(ws_main[f'E{row}'].value) for row in range(6, next_row + 1)], timedelta())
            total_count = sum([ws_main[f'C{row}'].value for row in range(6, next_row + 1)])
            avg_time = total_time / total_count if total_count > 0 else timedelta()

            ws_main['F2'] = str(avg_time)

            # Apply conditional formatting
            h2_value = str_to_timedelta(ws_main['H2'].value)
            if h2_value < avg_time:
                ws_main['H2'].fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            else:
                ws_main['H2'].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        wb.save(filename)
        print(f"Log updated: {filename}")

    except Exception as e:
        print(f"Error updating log: {e}")
        print(f"Data type: {type(data)}")
        print(f"Data content: {data}")
def finish_batch():
    global amount_of_cycles_done, total_elapsed_time, downtime, skipped_tests

    batch_date = datetime.now().strftime('%y-%m-%d %H:%M')
    
    script_dir = os.path.dirname(os.path.abspath(__file__))
    local_log_filepath = os.path.join(script_dir, "Artiklar", f"{filename}_log.xlsx")
    nas_log_filepath = f"/mnt/nas/Artiklar/{filename}_log.xlsx"

    # Log each cycle
    for i in range(amount_of_cycles_done):
        cycle_data = {
            "Batchdatum": batch_date,
            "Antal skippad test": skipped_tests,
            "Serienummer": i + 1
        }
        
        update_log(local_log_filepath, cycle_data, is_cycle_data=True)
        update_log(nas_log_filepath, cycle_data, is_cycle_data=True)

    # Log batch summary
    batch_data = {
        "Batchdatum": batch_date,
        "Antal": amount_of_cycles_done,
        "Antal skippad test": skipped_tests,
        "Total Cykeltid (HH:MM:SS)": seconds_to_hms(total_elapsed_time + downtime),
        "Total Ställtid (HH:MM:SS)": seconds_to_hms(downtime),
        "Total Stycktid (HH:MM:SS)": seconds_to_hms(total_elapsed_time),
        "Cykeltid (HH:MM:SS)": seconds_to_hms((total_elapsed_time + downtime) // amount_of_cycles_done if amount_of_cycles_done > 0 else 0),
        "Stycktid (HH:MM:SS)": seconds_to_hms(total_elapsed_time // amount_of_cycles_done if amount_of_cycles_done > 0 else 0),
        "Styck Ställtid (HH:MM:SS)": seconds_to_hms(downtime // amount_of_cycles_done if amount_of_cycles_done > 0 else 0)
    }
    
    update_log(local_log_filepath, batch_data)
    update_log(nas_log_filepath, batch_data)

    # Reset batch variables
    amount_of_cycles_done = 0
    total_elapsed_time = 0
    downtime = 0
    skipped_tests = 0

    # Update the labels
    completed_label.config(text=f"Färdiga: {amount_of_cycles_done}st")
    skipped_label.config(text=f"Antal Avvikande: {skipped_tests}st")

    print(f"Log files updated locally and on NAS for {filename}")
    # Reset serial number for the next batch
    initialize_serial_number()






# Header Section
header_frame = tk.Frame(root)
header_frame.pack(side=tk.TOP, fill=tk.X)


def exit_program(event=None):
    try:
        # Attempt to run finish_batch()
        finish_batch()
        print("Batch finished successfully.")
    except Exception as e:
        # If an error occurs, print it and continue with exit
        print(f"Error in finish_batch(): {e}")
        traceback.print_exc()  # This will print the full traceback for debugging
    finally:
        # Close the window and exit the program
        root.destroy()
        sys.exit()

# Open and resize the image
image = Image.open(logo_path)
scale_percent = 50  # percent of original size
width, height = image.size
new_width = int(width * scale_percent / 100)
new_height = int(height * scale_percent / 100)
resized_image = image.resize((new_width, new_height), Image.LANCZOS)
logo_image = ImageTk.PhotoImage(resized_image)

logo_label = tk.Label(header_frame, image=logo_image, cursor="hand2")  # Change cursor to indicate it's clickable
logo_label.pack(side=tk.LEFT, padx=10, pady=10)
logo_label.bind("<Button-1>", exit_program)  # Bind the new exit_program function


article_label = tk.Label(header_frame, text=f"Artikelnummer: {filename}", font=body_font)
article_label.pack(side=tk.LEFT, padx=10, pady=10)

completed_label = tk.Label(header_frame, text=f"Färdiga: {amount_of_cycles_done}st", font=body_font, bg="#32CD32", fg="Black")
completed_label.pack(side=tk.RIGHT, padx=10, pady=10)

skipped_label = tk.Label(header_frame, text=f"Antal Avvikande: {skipped_tests}st", font=body_font, bg="red", fg="Black")
skipped_label.pack(side=tk.RIGHT, padx=10, pady=10)

# Left Panel
left_panel = tk.Frame(root)
left_panel.pack(side=tk.LEFT, fill=tk.Y, padx=20, pady=0)

left_panel_labels = []
for i, pin in enumerate(pins):
    label = tk.Label(left_panel, text=pin, font=body_font, justify=tk.LEFT, width=8, height=2)
    label.bind("<Button-1>", lambda e, idx=i: on_pin_click(idx))
    left_panel_labels.append(label)
    row = i % 16
    col = i // 16
    label.grid(row=row, column=col, sticky='w')


# Right Panel
right_panel = tk.Frame(root)
right_panel.pack(side=tk.RIGHT, fill=tk.Y, padx=20, pady=50)


time_info_label = tk.Label(right_panel, text=f"Tid\nNu: {format_time(elapsed_time_current_cycle)}\nFörra: {format_time(elapsed_time_previous_cycle)}\nTotal: {format_time(total_elapsed_time)}\nStälltid: {format_time(downtime)}", font=body_font, justify=tk.LEFT)
time_info_label.pack()

# Central Display
central_frame = tk.Frame(root)
central_frame.pack(expand=True, padx=0)

current_wire_label = tk.Label(central_frame, text="Starta", font=center_font, bg="#32CD32", width=6, height=3)
current_wire_label.pack(pady=20)
current_wire_label.bind("<Button-1>", lambda e: toggle_timer())

# Buttons
button_frame = tk.Frame(root)
button_frame.pack(side=tk.BOTTOM, fill=tk.X, pady=20)

# Button Frame
button_frame = tk.Frame(root)
button_frame.pack(side=tk.BOTTOM, fill=tk.X, pady=10)



# Update the diagnose button to finish batch
reset_button = tk.Button(button_frame, text="Reset", font=("Helvetica", 24), bg="#9900AB", fg="black", command=reset_test, width=15, height=50)
reset_button.pack(side=tk.LEFT, padx=5, pady=10)

# Button for manually controlling the relay
manual_probe_button = tk.Button(button_frame, text="Control Relay", font=("Helvetica", 24), bg="#FFA500", fg="black", command=manual_relay_control, width=15, height=50)
manual_probe_button.pack(side=tk.LEFT, padx=5, pady=10)

finish_batch_button = tk.Button(button_frame, text="Finish Batch", font=("Helvetica", 24), bg="#0A60C5", fg="black", command=finish_batch, width=15, height=50)
finish_batch_button.pack(side=tk.RIGHT, padx=5, pady=10)
# Create the motor control button in your UI setup
motor_button = tk.Button(button_frame, text="Motor", font=("Helvetica", 16), command=run_motor, bg="gray", width=25, height=50)
motor_button.pack(side=tk.RIGHT, padx=5, pady=10)




# Start the timer
root.after(1000, update_timer)  # Start the timer after the first second

# Initialize the first pin
left_panel_labels[current_pin_index].config(bg="yellow")

# Run the main loop
root.mainloop()
